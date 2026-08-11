import io
import uuid
import hashlib
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from openpyxl import load_workbook

app = FastAPI(title="Tax Prep")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory sessions keyed by email hash
sessions: dict = {}

CATEGORIES_FILE = Path(__file__).parent / "CATEGORIES.txt"


# ── helpers ────────────────────────────────────────────────────────────────────

def session_key(email: str) -> str:
    return hashlib.sha256(email.strip().lower().encode()).hexdigest()


def load_categories(text: str) -> dict:
    cats = {}
    for line in text.splitlines():
        line = line.strip()
        if line and ":" in line:
            key, val = line.split(":", 1)
            cats[key.strip().lower()] = val.strip()
    return cats


def categorize(description: str, lookup: dict) -> str:
    dl = description.lower()
    for kw, cat in lookup.items():
        if kw in dl:
            return cat
    return "Uncategorized"


def parse_date(s) -> str:
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(s).strip(), fmt).strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            continue
    return str(s)


def safe_float(v) -> float:
    try:
        return float(str(v).replace(",", "").replace("$", "")) if v and str(v).strip() not in ("", "nan") else 0.0
    except Exception:
        return 0.0


def get_session(sid: str) -> dict:
    if sid not in sessions:
        raise HTTPException(status_code=404, detail="Session not found. Please log in again.")
    return sessions[sid]


def build_summary(txs: list) -> list:
    summary: dict = {}
    for tx in txs:
        cat = tx["category"]
        if cat not in summary:
            summary[cat] = {"category": cat, "out": 0.0, "in_amount": 0.0, "count": 0}
        summary[cat]["out"] += tx["out"]
        summary[cat]["in_amount"] += tx["in_amount"]
        summary[cat]["count"] += 1
    result = sorted(summary.values(), key=lambda x: x["category"])
    result.append({
        "category": "TOTAL",
        "out": sum(r["out"] for r in result),
        "in_amount": sum(r["in_amount"] for r in result),
        "count": sum(r["count"] for r in result),
    })
    return result


# ── routes ────────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def home():
    return (Path(__file__).parent / "static" / "index.html").read_text()


@app.post("/api/login")
async def login(name: str = Form(...), email: str = Form(...)):
    """Create or resume a session keyed by email."""
    sid = session_key(email)
    if sid not in sessions:
        cat_text = CATEGORIES_FILE.read_text() if CATEGORIES_FILE.exists() else ""
        sessions[sid] = {
            "name": name,
            "email": email.strip().lower(),
            "transactions": [],
            "categories": load_categories(cat_text),
        }
    else:
        # Update name in case it changed
        sessions[sid]["name"] = name
    return {"session_id": sid, "name": sessions[sid]["name"], "email": sessions[sid]["email"],
            "transaction_count": len(sessions[sid]["transactions"])}


@app.post("/api/session/{sid}/upload")
async def upload(sid: str, source: str = Form(...), files: List[UploadFile] = File(...)):
    sess = get_session(sid)
    lookup = sess["categories"]
    added = 0

    for file in files:
        content = await file.read()
        try:
            df = pd.read_csv(io.BytesIO(content), header=None, dtype=str)
            if df.shape[1] < 4:
                continue
            df = df.iloc[:, :5]
            df.columns = ["DATE", "DESCRIPTION", "OUT", "IN", "BALANCE"][: df.shape[1]]

            for _, row in df.iterrows():
                desc = str(row.get("DESCRIPTION", "") or "").strip()
                if not desc:
                    continue
                tx = {
                    "id": str(uuid.uuid4()),
                    "date": parse_date(row.get("DATE", "")),
                    "description": desc,
                    "out": safe_float(row.get("OUT")),
                    "in_amount": safe_float(row.get("IN")),
                    "source": source.upper(),
                    "category": categorize(desc, lookup),
                    "file": file.filename,
                }
                sess["transactions"].append(tx)
                added += 1
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading {file.filename}: {e}")

    # Deduplicate
    seen: set = set()
    deduped = []
    for tx in sess["transactions"]:
        key = (tx["date"], tx["description"], tx["out"], tx["in_amount"])
        if key not in seen:
            seen.add(key)
            deduped.append(tx)
    removed = len(sess["transactions"]) - len(deduped)
    sess["transactions"] = deduped

    return {"added": added, "duplicates_removed": removed, "total": len(sess["transactions"])}


@app.get("/api/session/{sid}/transactions")
async def get_transactions(sid: str, category: Optional[str] = None, search: Optional[str] = None):
    sess = get_session(sid)
    txs = sess["transactions"]
    if category and category != "all":
        txs = [t for t in txs if t["category"] == category]
    if search:
        sl = search.lower()
        txs = [t for t in txs if sl in t["description"].lower()]
    return sorted(txs, key=lambda x: (x["category"], x["date"], x["description"]))


@app.patch("/api/session/{sid}/transactions/{tx_id}")
async def update_transaction(sid: str, tx_id: str, body: dict):
    sess = get_session(sid)
    for tx in sess["transactions"]:
        if tx["id"] == tx_id:
            if "category" in body:
                tx["category"] = body["category"]
            return tx
    raise HTTPException(status_code=404, detail="Transaction not found")


@app.get("/api/session/{sid}/summary")
async def get_summary(sid: str):
    sess = get_session(sid)
    return build_summary(sess["transactions"])


@app.get("/api/session/{sid}/category-list")
async def get_category_list(sid: str):
    """All unique category names — from lookup + current transactions."""
    sess = get_session(sid)
    from_lookup = set(sess["categories"].values())
    from_txs = set(tx["category"] for tx in sess["transactions"])
    return sorted(from_lookup | from_txs)


@app.delete("/api/session/{sid}/transactions")
async def clear_transactions(sid: str):
    sess = get_session(sid)
    sess["transactions"] = []
    return {"message": "Cleared"}


@app.get("/api/session/{sid}/export")
async def export_excel(sid: str):
    sess = get_session(sid)
    txs = sess["transactions"]
    if not txs:
        raise HTTPException(status_code=400, detail="No transactions to export.")

    df = pd.DataFrame(txs)[["date", "description", "out", "in_amount", "category", "source"]]
    df.columns = ["DATE", "DESCRIPTION", "OUT", "IN", "CATEGORY", "SOURCE"]
    df = df.sort_values(["CATEGORY", "DESCRIPTION"])

    summary = {}
    for tx in txs:
        cat = tx["category"]
        if cat not in summary:
            summary[cat] = {"CATEGORY": cat, "OUT": 0.0, "IN": 0.0}
        summary[cat]["OUT"] += tx["out"]
        summary[cat]["IN"] += tx["in_amount"]
    summary_df = pd.DataFrame(sorted(summary.values(), key=lambda x: x["CATEGORY"]))
    total_row = pd.DataFrame([{"CATEGORY": "TOTAL", "OUT": summary_df["OUT"].sum(), "IN": summary_df["IN"].sum()}])
    summary_df = pd.concat([summary_df, total_row], ignore_index=True)

    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    fp = f"/tmp/TAXES_MASTER_{ts}.xlsx"

    with pd.ExcelWriter(fp, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Transactions", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    wb = load_workbook(fp)
    for sh in wb.sheetnames:
        ws = wb[sh]
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 2, 60)
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
            for cell in row:
                cell.number_format = '"$"#,##0.00'
    wb.save(fp)

    return FileResponse(
        fp,
        filename=f"TAXES_MASTER_{ts}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
